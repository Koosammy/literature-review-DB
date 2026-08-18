import io
import re
from datetime import datetime, timedelta
from typing import List

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, UploadFile, File, status
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError

from ..database import get_db
from ..models.user import User
from ..schemas.user import UserCreate, UserUpdate, UserResponse
from ..core.auth import get_current_active_user, require_main_coordinator
from ..core.security import get_password_hash
from ..core.config import settings
from ..core.constants import SCHOOLS
from ..core.email import generate_reset_token, send_welcome_email

router = APIRouter()

INVITE_EXPIRY = timedelta(days=7)


def _generate_username(db: Session, full_name: str, email: str) -> str:
    """Derive a username from the email's local part (falling back to the
    name) since neither the "Add User" form nor the Excel import collects
    one anymore. Deduplicated with a numeric suffix if already taken."""
    base = re.sub(r'[^a-z0-9_-]', '', email.split('@')[0].lower())
    if len(base) < 3:
        base = re.sub(r'[^a-z0-9_-]', '', full_name.lower().replace(' ', '_'))
    if len(base) < 3:
        base = f"user_{base}"
    base = base[:40]

    username = base
    suffix = 1
    while db.query(User).filter(User.username == username).first():
        suffix += 1
        username = f"{base}{suffix}"
    return username


def _normalize_school(raw: str) -> str:
    raw = (raw or "").strip()
    if not raw:
        return raw
    for school in SCHOOLS:
        if raw.lower() == school.lower() or raw.lower() == school.split(" - ", 1)[-1].lower():
            return school
    return raw


def _invite_user(
    db: Session,
    *,
    email: str,
    full_name: str,
    title: str = None,
    institution: str = None,
    department: str = None,
    role: str = "faculty",
) -> User:
    """Create a user with no usable password yet and a pending activation
    token. Caller is responsible for committing and sending the email."""
    placeholder = get_password_hash(generate_reset_token(48))
    user = User(
        username=_generate_username(db, full_name, email),
        email=email,
        hashed_password=placeholder,
        title=title,
        full_name=full_name,
        institution=institution,
        department=department,
        role=role,
        is_active=True,
        must_set_password=True,
        reset_token=generate_reset_token(),
        reset_token_expires=datetime.utcnow() + INVITE_EXPIRY,
    )
    db.add(user)
    return user


def _activation_url(token: str) -> str:
    return f"{settings.FRONTEND_URL}/#/activate-account?token={token}"


@router.get("/", response_model=List[UserResponse])
async def get_users(
    skip: int = 0,
    limit: int = 100,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    users = db.query(User).offset(skip).limit(limit).all()
    return users

@router.post("/", response_model=UserResponse)
async def create_user(
    user: UserCreate,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(require_main_coordinator),
    db: Session = Depends(get_db)
):
    """Create a user and email them an activation link. Admins no longer
    choose a password here -- the user sets their own (and uploads their
    mandatory profile photo) via the link before they can sign in."""
    try:
        if db.query(User).filter(User.email == user.email).first():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Email already exists"
            )

        db_user = _invite_user(
            db,
            email=user.email,
            full_name=user.full_name,
            title=user.title,
            institution=_normalize_school(user.institution) if user.institution else None,
            department=user.department,
            role=user.role,
        )

        db.commit()
        db.refresh(db_user)

        background_tasks.add_task(
            send_welcome_email,
            email=db_user.email,
            full_name=db_user.full_name,
            activation_url=_activation_url(db_user.reset_token),
        )

        return db_user

    except HTTPException:
        raise
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Database constraint violation. Email may already exist."
        )
    except Exception:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An error occurred while creating the user"
        )


@router.post("/bulk-import")
async def bulk_import_users(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    current_user: User = Depends(require_main_coordinator),
    db: Session = Depends(get_db)
):
    """Create users in bulk from an uploaded .xlsx with columns
    Title, Full Name, Email, School. Each created user gets the same
    activation-email flow as a single manually-created user. Rows whose
    email already belongs to an existing user are skipped (not
    overwritten) and reported back, per row, in the response."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel import is unavailable on this server (openpyxl not installed)."
        )

    filename = (file.filename or "").lower()
    if not filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Please upload a .xlsx file."
        )

    raw = await file.read()
    try:
        workbook = load_workbook(io.BytesIO(raw), data_only=True)
        sheet = workbook.active
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Could not read the uploaded file. Please make sure it's a valid .xlsx file."
        )

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="The file is empty.")

    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    required = ["title", "full name", "email", "school"]
    missing = [col for col in required if col not in header]
    if missing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Missing required column(s): {', '.join(missing)}. "
                   f"Expected columns: Title, Full Name, Email, School."
        )
    col_index = {name: header.index(name) for name in required}

    created: List[dict] = []
    skipped: List[dict] = []
    seen_emails_this_file = set()

    for row_num, row in enumerate(rows[1:], start=2):
        def cell(name: str) -> str:
            idx = col_index[name]
            value = row[idx] if idx < len(row) else None
            return str(value).strip() if value is not None else ""

        email = cell("email").lower()
        full_name = cell("full name")
        title = cell("title") or None
        school = cell("school")

        if not email and not full_name:
            continue  # blank row

        if not email or not full_name:
            skipped.append({"row": row_num, "email": email or None, "reason": "Missing required field(s)"})
            continue

        if not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email):
            skipped.append({"row": row_num, "email": email, "reason": "Invalid email address"})
            continue

        if email in seen_emails_this_file:
            skipped.append({"row": row_num, "email": email, "reason": "Duplicate email within this file"})
            continue

        if db.query(User).filter(User.email == email).first():
            skipped.append({"row": row_num, "email": email, "reason": "A user with this email already exists"})
            continue

        db_user = _invite_user(
            db,
            email=email,
            full_name=full_name,
            title=title,
            institution=_normalize_school(school) if school else None,
            role="faculty",
        )
        try:
            db.flush()
        except IntegrityError:
            db.rollback()
            skipped.append({"row": row_num, "email": email, "reason": "A user with this email already exists"})
            continue

        seen_emails_this_file.add(email)
        created.append({"row": row_num, "email": email, "full_name": full_name, "user_id": db_user.id})

    db.commit()

    for entry in created:
        db_user = db.query(User).filter(User.id == entry["user_id"]).first()
        if db_user:
            background_tasks.add_task(
                send_welcome_email,
                email=db_user.email,
                full_name=db_user.full_name,
                activation_url=_activation_url(db_user.reset_token),
            )

    return {
        "created_count": len(created),
        "skipped_count": len(skipped),
        "created": created,
        "skipped": skipped,
    }


@router.post("/{user_id}/resend-invite")
async def resend_invite(
    user_id: int,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(require_main_coordinator),
    db: Session = Depends(get_db)
):
    """Re-send the activation email with a fresh token (e.g. the original
    link expired, or landed in spam)."""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    if not user.must_set_password:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="This user has already activated their account."
        )

    user.reset_token = generate_reset_token()
    user.reset_token_expires = datetime.utcnow() + INVITE_EXPIRY
    db.commit()

    background_tasks.add_task(
        send_welcome_email,
        email=user.email,
        full_name=user.full_name,
        activation_url=_activation_url(user.reset_token),
        is_resend=True,
    )

    return {"message": f"Activation email resent to {user.email}"}

@router.get("/{user_id}", response_model=UserResponse)
async def get_user(
    user_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found"
        )
    return user

@router.put("/{user_id}", response_model=UserResponse)
async def update_user(
    user_id: int,
    user_update: UserUpdate,
    current_user: User = Depends(require_main_coordinator),
    db: Session = Depends(get_db)
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found"
        )

    # Update user fields
    update_data = user_update.dict(exclude_unset=True)
    if "institution" in update_data and update_data["institution"]:
        update_data["institution"] = _normalize_school(update_data["institution"])
    for field, value in update_data.items():
        setattr(user, field, value)

    try:
        db.commit()
        db.refresh(user)
        return user
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Email already exists"
        )

@router.delete("/{user_id}")
async def delete_user(
    user_id: int,
    current_user: User = Depends(require_main_coordinator),
    db: Session = Depends(get_db)
):
    # Prevent deleting yourself
    if user_id == current_user.id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot delete your own account"
        )

    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found"
        )

    try:
        db.delete(user)
        db.commit()
        return {"message": "User deleted successfully"}
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An error occurred while deleting the user"
        )

@router.patch("/{user_id}/toggle-status")
async def toggle_user_status(
    user_id: int,
    current_user: User = Depends(require_main_coordinator),
    db: Session = Depends(get_db)
):
    # Prevent deactivating yourself
    if user_id == current_user.id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot deactivate your own account"
        )

    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found"
        )

    try:
        user.is_active = not user.is_active
        db.commit()
        db.refresh(user)

        status_text = "activated" if user.is_active else "deactivated"
        return {
            "message": f"User {status_text} successfully",
            "user": {
                "id": user.id,
                "username": user.username,
                "is_active": user.is_active
            }
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An error occurred while updating user status"
        )
